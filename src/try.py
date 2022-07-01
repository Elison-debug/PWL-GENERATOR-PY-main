from ast import Str
from heapq import nsmallest
from openpyxl import load_workbook
from engineering_notation import EngNumber
unit="ns"
Step_length=100
Excel_name="text.xlsx"
result_name="result.txt"

#voltage_level = 5
# risetime=10
# falltime=20

wb = load_workbook(filename=Excel_name)
with open (result_name,"w+") as f: 
    f.write("designed by Elison\n")
#f = open("my_file.txt", 'w+',encoding = "utf-8")




#f.close()

sheets = wb.sheetnames
#第一个表格的名称
sheet_first = sheets[0]
#获取特定的worksheet
ws = wb[sheet_first]
 
#获取表格所有行和列，两者都是可迭代的
rows = list(ws.rows)
columns = list(ws.columns)

#迭代所有的行，输出PWL
for row in rows[2:len(rows)]:

 with open (result_name,"a+") as f:
    if row[0].value!=None:
        f.write("v"+str(row[0].value)+" PWL ")
    else :
        break

    for col in row[1:5]:
        if type(col.value)==str:
            f.write(str(col.value)+" ")

    voltage_level_high=row[1].value if float(row[1].value)>float(row[2].value) else row[2].value
    voltage_level_low =row[1].value if float(row[1].value)<float(row[2].value) else row[2].value
    if row[1]==voltage_level_low:
        state=0
    else :
        state=1
    risetime=int(row[3].value)
    falltime=int(row[4].value)
    for col in row[5:len(row)]:
      if type(col.value)==str:
        float(col.value)
        if state==0:
            f.write(str(col.value)+" "+voltage_level_low +" "+str(float(col.value)+float(risetime))+" "+voltage_level_high+" ")
            state=1
        else :
            f.write(str(col.value)+" "+voltage_level_high+" "+str(float(col.value)+float(falltime))+" "+voltage_level_low +" ")
            state=0

      else :
        f.write("              ")
    f.write("\n")

      
#通过坐标读取值
#print (ws.cell('A1').value) # A表示列,1表示行
print (ws.cell(row=1, column=1).value)


# def make_pwl(sequence:str,bit_rate:float,filename="pwl.txt",voltage_level = 5):
#     """Python Function to generate digital input voltage sequence
#     given bit sequence and datarate. Output waveform
#     at 5V voltage level by default and saved to a
#     text file.

#     Args:
#         sequence (str): Sequence of bits ex: "1011001"
#         bit_rate (float): data rate in Hz ex: 1e3 for 1 Kbps
#         filename (str, optional): File to save the generated PWL. Defaults to "pwl.txt".
#         voltage_level (int, optional): Peak Voltage level in PWL. Defaults to 5.
#     """ 
#     bit_period = 1/bit_rate
#     half = bit_period/2
#     mid = bit_period/2
#     with open(filename,"w") as f:
#         for bit in sequence:
#             f.write(f"{EngNumber(mid - half,5)}\t{int(bit) * voltage_level}\n")
#             f.write(f"{EngNumber(mid + half - half/100000,5 )}\t{int(bit) * voltage_level}\n")
#             mid = bit_period + mid
            
#     f.close()
    
    
# if __name__ == "__main__":
#     import pandas as pd
#     # Example Sequence
#     sequence = "0111100"
#     frequency = 5
#     print(make_pwl(sequence,frequency))
    
