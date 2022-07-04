from openpyxl import load_workbook
#单位
unit="ns"
#步进
Step_length=100
#保存文件名
result_name="result/result.txt"
voltage_level = 5



# risetime=10
# falltime=20

wb = load_workbook(filename="src/test.xlsx")
with open (result_name,"w+") as f: 
    f.write("design by Elison\n")
#f = open("my_file.txt", 'w+',encoding = "utf-8")




#f.close()

sheets = wb.sheetnames
#第一个表格的名称
sheet_first = sheets[0]
#获取特定的worksheet
ws = wb[sheet_first]
 
#获取表格所有行和列，两者都是可迭代的
rows = list(ws.rows)
columns = list(ws.columns)

#迭代所有的行，输出PWL
for row in rows[2:len(rows)]:

 with open (result_name,"a+") as f:
    if row[0].value!=None:
        f.write("v"+str(row[0].value)+" PWL ")
    else :
        break



    voltage_level_high=row[1].value if float(row[1].value)>float(row[2].value) else row[2].value
    voltage_level_low =row[1].value if float(row[1].value)<float(row[2].value) else row[2].value
    if row[1]==voltage_level_low:
        state=0
    else :
        state=1
    risetime=int(row[3].value)
    falltime=int(row[4].value)
    for col in row[5:len(row)]:
      if type(col.value)==str:
        float(col.value)
        if state==0:
            f.write(str(col.value)+" "+voltage_level_low +" "+str(float(col.value)+float(risetime))+" "+voltage_level_high+" ")
            state=1
        else :
            f.write(str(col.value)+" "+voltage_level_high+" "+str(float(col.value)+float(falltime))+" "+voltage_level_low +" ")
            state=0

      else :
        f.write("              ")
    f.write("\n")